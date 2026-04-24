import pandas as pd
import openpyxl
import os
import sys

def process_excel(file_path, output_dir):
    print(f"Processing {file_path}...")
    # data_only=True to get values instead of formulas
    wb = openpyxl.load_workbook(file_path, data_only=True)
    
    file_name = os.path.basename(file_path).replace('.xlsx', '')
    
    for sheet_name in wb.sheetnames:
        print(f"  Sheet: {sheet_name}")
        ws = wb[sheet_name]
        
        # Handle merged cells: store value, unmerge, and fill
        merged_ranges = list(ws.merged_cells.ranges)
        for merged_range in merged_ranges:
            min_col, min_row, max_col, max_row = merged_range.bounds
            top_left_value = ws.cell(row=min_row, column=min_col).value
            
            # Unmerge before writing to the constituent cells
            ws.unmerge_cells(str(merged_range))
            
            for row in range(min_row, max_row + 1):
                for col in range(min_col, max_col + 1):
                    ws.cell(row=row, column=col).value = top_left_value
        
        # Convert to DataFrame
        # ws.values returns a generator of tuples
        data = list(ws.values)
        if not data:
            print(f"    Sheet {sheet_name} is empty.")
            continue
            
        # Use first row as columns initially
        df = pd.DataFrame(data)
        
        # Basic cleaning: Drop rows and columns that are entirely NaN
        df.dropna(how='all', axis=0, inplace=True)
        df.dropna(how='all', axis=1, inplace=True)
        
        # Reset index
        df.reset_index(drop=True, inplace=True)
        
        # Sanitize sheet name for file system
        safe_sheet_name = "".join([c if c.isalnum() else "_" for c in sheet_name])
        output_file = os.path.join(output_dir, f"{file_name}_{safe_sheet_name}.csv")
        
        df.to_csv(output_file, index=False, header=False) # No header yet as it might be messy
        print(f"    Saved to {output_file}")

if __name__ == "__main__":
    output_dir = "csv_output"
    os.makedirs(output_dir, exist_ok=True)
    
    files = [f for f in os.listdir('.') if f.endswith('.xlsx')]
    if not files:
        print("No .xlsx files found in current directory.")
        sys.exit(0)
        
    for f in files:
        process_excel(f, output_dir)
